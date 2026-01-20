from flask import Blueprint, render_template, request, redirect, url_for, session, flash, jsonify
import datetime
from ..extensions import db
from ..models import PaymentRequest, Evidence, Estado, Sociedad
from ..services.telegram import send_message
from sqlalchemy import or_, func

admin_bp = Blueprint("admin_bp", __name__)


def require_admin(f):
    from functools import wraps

    def _is_admin():
        return session.get("is_admin")

    @wraps(f)
    def wrapper(*a, **k):
        if _is_admin():
            return f(*a, **k)
        return redirect(url_for("admin_bp.login"))

    return wrapper


@admin_bp.get("/login")
def login():
    return render_template("login.html")


@admin_bp.post("/login")
def do_login():
    from flask import current_app

    if request.form.get("password") == current_app.config["ADMIN_PASSWORD"]:
        session["is_admin"] = True
        return redirect(url_for("admin_bp.admin"))
    flash("Contraseña incorrecta", "danger")
    return redirect(url_for("admin_bp.login"))


@admin_bp.post("/logout")
@require_admin
def logout():
    session.clear()
    return redirect(url_for("admin_bp.login"))


@admin_bp.get("/")
def root():
    return redirect(
        url_for("admin_bp.admin")
        if session.get("is_admin")
        else url_for("admin_bp.login")
    )


@admin_bp.get("/admin")
@require_admin
def admin():
    estado = request.args.get("estado", "").strip()
    q_str = request.args.get("q", "").strip()
    desde_str = request.args.get("desde", "").strip()
    hasta_str = request.args.get("hasta", "").strip()
    sociedad_str = request.args.get("sociedad", "").strip().upper()
    valor_min_str = request.args.get("valor_min", "").strip()
    valor_max_str = request.args.get("valor_max", "").strip()
    page = int(request.args.get("page", 1))
    per_page = int(request.args.get("per_page", 25))

    vmin = None
    vmax = None
    if valor_min_str:
        try:
            vmin = int(valor_min_str)
        except Exception:
            vmin = None
    if valor_max_str:
        try:
            vmax = int(valor_max_str)
        except Exception:
            vmax = None
    if vmin is not None and vmax is not None and vmin > vmax:
        vmin, vmax = vmax, vmin

    # Base para totales por estado (no filtra 'estado' para ver distribución completa)
    base = PaymentRequest.query
    if q_str:
        like = f"%{q_str}%"
        base = base.filter(
            or_(
                PaymentRequest.cliente.like(like),
                PaymentRequest.medio_pago.like(like),
                PaymentRequest.sucursal.like(like),
            )
        )
    if desde_str:
        try:
            d = datetime.datetime.strptime(desde_str, "%Y-%m-%d")
            base = base.filter(PaymentRequest.created_at >= d)
        except Exception:
            pass
    if hasta_str:
        try:
            h = datetime.datetime.strptime(hasta_str, "%Y-%m-%d") + datetime.timedelta(days=1)
            base = base.filter(PaymentRequest.created_at < h)
        except Exception:
            pass
    # Filtro sociedad (en agregados)
    soc_val = None
    if sociedad_str in [s.value for s in Sociedad]:
        soc_val = Sociedad(sociedad_str)
        base = base.filter(PaymentRequest.sociedad == soc_val)
    if vmin is not None:
        base = base.filter(PaymentRequest.valor >= vmin)
    if vmax is not None:
        base = base.filter(PaymentRequest.valor <= vmax)

    # Conteo por estado
    counts_raw = (
        base.with_entities(PaymentRequest.estado, func.count())
        .group_by(PaymentRequest.estado)
        .all()
    )
    counts_by_status = {e.value: 0 for e in Estado}
    for est, cnt in counts_raw:
        try:
            key = est.value if est else ""
        except Exception:
            key = str(est) if est else ""
        if key:
            counts_by_status[key] = int(cnt)

    # Suma total y por estado (valor)
    sum_total = base.with_entities(func.sum(PaymentRequest.valor)).scalar() or 0
    sums_raw = (
        base.with_entities(PaymentRequest.estado, func.sum(PaymentRequest.valor))
        .group_by(PaymentRequest.estado)
        .all()
    )
    sums_by_status = {e.value: 0 for e in Estado}
    for est, s in sums_raw:
        try:
            key = est.value if est else ""
        except Exception:
            key = str(est) if est else ""
        if key:
            sums_by_status[key] = int(s or 0)

    # Query principal para listado
    query = PaymentRequest.query
    if estado in [e.value for e in Estado]:
        query = query.filter(PaymentRequest.estado == Estado(estado))
    if q_str:
        like = f"%{q_str}%"
        query = query.filter(
            or_(
                PaymentRequest.cliente.like(like),
                PaymentRequest.medio_pago.like(like),
                PaymentRequest.sucursal.like(like),
            )
        )
    # Filtros de fecha (UTC) en created_at
    if desde_str:
        try:
            d = datetime.datetime.strptime(desde_str, "%Y-%m-%d")
            query = query.filter(PaymentRequest.created_at >= d)
        except Exception:
            pass
    if hasta_str:
        try:
            h = datetime.datetime.strptime(hasta_str, "%Y-%m-%d") + datetime.timedelta(days=1)
            query = query.filter(PaymentRequest.created_at < h)
        except Exception:
            pass
    if soc_val is not None:
        query = query.filter(PaymentRequest.sociedad == soc_val)
    if vmin is not None:
        query = query.filter(PaymentRequest.valor >= vmin)
    if vmax is not None:
        query = query.filter(PaymentRequest.valor <= vmax)

    total = query.count()
    pagos = (
        query.order_by(PaymentRequest.created_at.desc())
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )

    return render_template(
        "admin.html",
        pagos=pagos,
        Estado=Estado,
        page=page,
        per_page=per_page,
        total=total,
        estado=estado,
        q=q_str,
        desde=desde_str,
        hasta=hasta_str,
        sociedad=sociedad_str,
        counts_by_status=counts_by_status,
        sum_total=sum_total,
        sums_by_status=sums_by_status,
        valor_min=valor_min_str,
        valor_max=valor_max_str,
    )


@admin_bp.get("/evidence/<int:evid_id>")
@require_admin
def evidence_view(evid_id):
    from flask import current_app, send_from_directory

    ev = Evidence.query.get_or_404(evid_id)
    return send_from_directory(
        current_app.config["EVID_DIR"], ev.filename, as_attachment=False
    )


@admin_bp.post("/payments/<int:pid>/approve")
@require_admin
def approve(pid):
    p = PaymentRequest.query.get_or_404(pid)
    if p.estado == Estado.PENDIENTE:
        p.estado = Estado.APROBADO
        db.session.commit()
        send_message(
            p.chat_id_respuesta,
            f"✅ Pago de <b>{p.cliente}</b> fue <b>APROBADO</b>.\nID: <b>{p.id}</b> | Valor: ${p.valor:,}",
        )
    return redirect(url_for("admin_bp.admin"))


@admin_bp.post("/payments/<int:pid>/reject")
@require_admin
def reject(pid):
    p = PaymentRequest.query.get_or_404(pid)
    if p.estado == Estado.PENDIENTE:
        motivo = (request.form.get("motivo") or "No cumple validación").strip()
        p.estado = Estado.RECHAZADO
        p.motivo_rechazo = motivo
        db.session.commit()
        send_message(
            p.chat_id_respuesta,
            f"❌ Pago de <b>{p.cliente}</b> fue <b>RECHAZADO</b>.\nMotivo: {motivo}\nID: <b>{p.id}</b>",
        )
    return redirect(url_for("admin_bp.admin"))


@admin_bp.get("/health")
def health():
    return jsonify({"status": "ok"})


# --- EXPORTAR BANDEJA A EXCEL ---
@admin_bp.get("/payments/export-excel")
@require_admin
def export_payments_excel():
    import os
    from io import BytesIO
    from flask import send_file, current_app
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage

    include_images = request.args.get("imagenes", "").strip().lower() in [
        "1",
        "true",
        "t",
        "yes",
        "si",
        "sí",
    ]
    estado = request.args.get("estado")
    q_str = (request.args.get("q") or "").strip()
    desde_str = (request.args.get("desde") or "").strip()
    hasta_str = (request.args.get("hasta") or "").strip()
    sociedad_str = (request.args.get("sociedad") or "").strip().upper()
    valor_min_str = (request.args.get("valor_min") or "").strip()
    valor_max_str = (request.args.get("valor_max") or "").strip()

    q = PaymentRequest.query
    if estado in [e.value for e in Estado]:
        q = q.filter(PaymentRequest.estado == Estado(estado))
    if q_str:
        like = f"%{q_str}%"
        q = q.filter(
            or_(
                PaymentRequest.cliente.like(like),
                PaymentRequest.medio_pago.like(like),
                PaymentRequest.sucursal.like(like),
            )
        )
    if desde_str:
        try:
            d = datetime.datetime.strptime(desde_str, "%Y-%m-%d")
            q = q.filter(PaymentRequest.created_at >= d)
        except Exception:
            pass
    if hasta_str:
        try:
            h = datetime.datetime.strptime(hasta_str, "%Y-%m-%d") + datetime.timedelta(days=1)
            q = q.filter(PaymentRequest.created_at < h)
        except Exception:
            pass
    if sociedad_str in [s.value for s in Sociedad]:
        q = q.filter(PaymentRequest.sociedad == Sociedad(sociedad_str))
    # Filtro por valor
    try:
        vmin = int(valor_min_str) if valor_min_str else None
    except Exception:
        vmin = None
    try:
        vmax = int(valor_max_str) if valor_max_str else None
    except Exception:
        vmax = None
    if vmin is not None and vmax is not None and vmin > vmax:
        vmin, vmax = vmax, vmin
    if vmin is not None:
        q = q.filter(PaymentRequest.valor >= vmin)
    if vmax is not None:
        q = q.filter(PaymentRequest.valor <= vmax)
    pagos = q.order_by(PaymentRequest.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Pagos"

    headers = [
        "ID",
        "Cliente",
        "Valor",
        "Medio de pago",
        "Sucursal",
        "Fecha consignación",
        "Sociedad",
        "Estado",
        "Motivo rechazo",
        "Creado UTC",
        "Actualizado UTC",
        "Telegram User ID",
        "Chat ID respuesta",
        "Evidencias (archivos)",
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True)
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    for p in pagos:
        evids = ", ".join(ev.filename or "" for ev in (p.evidences or []))
        ws.append(
            [
                p.id,
                getattr(p, "cliente", None)
                or getattr(
                    p, "referencia", ""
                ),
                p.valor or 0,
                p.medio_pago or "",
                p.sucursal or "",
                (p.fecha_consignacion.isoformat() if getattr(p, "fecha_consignacion", None) else ""),
                (p.sociedad.value if getattr(p, "sociedad", None) else ""),
                p.estado.value if p.estado else "",
                p.motivo_rechazo or "",
                (p.created_at.isoformat(sep=" ") if p.created_at else ""),
                (p.updated_at.isoformat(sep=" ") if p.updated_at else ""),
                p.telegram_user_id or "",
                p.chat_id_respuesta or "",
                evids,
            ]
        )

    # formato numérico (columna Valor = 3)
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = "#,##0"

    widths = [8, 28, 14, 22, 22, 18, 14, 12, 28, 20, 20, 16, 16, 36]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ----- Hoja con imágenes opcional -----
    if include_images:
        ws2 = wb.create_sheet(title="Evidencias")
        ws2.append(["Pago ID", "Cliente", "Archivo", "Imagen"])
        for col in range(1, 5):
            c = ws2.cell(row=1, column=col)
            c.font = Font(bold=True)
            c.alignment = Alignment(vertical="center")
        ws2.freeze_panes = "A2"
        ws2.column_dimensions["A"].width = 10
        ws2.column_dimensions["B"].width = 28
        ws2.column_dimensions["C"].width = 36
        ws2.column_dimensions["D"].width = 50  # donde anclamos la imagen

        # límites para thumbnails (px)
        MAX_W, MAX_H = 420, 300
        IMAGE_EXTS = {".jpg", ".jpeg", ".png"}

        row_idx = 2
        for p in pagos:
            cliente = getattr(p, "cliente", None) or getattr(p, "referencia", "")
            for ev in p.evidences or []:
                filename = ev.filename or ""
                ext = os.path.splitext(filename)[1].lower()
                img_path = os.path.join(current_app.config["EVID_DIR"], filename)
                ws2.cell(row=row_idx, column=1, value=p.id)
                ws2.cell(row=row_idx, column=2, value=cliente)
                ws2.cell(row=row_idx, column=3, value=filename)

                if ext in IMAGE_EXTS and os.path.exists(img_path):
                    try:
                        # Cargamos imagen y ajustamos tamaño
                        from PIL import Image as PILImage

                        with PILImage.open(img_path) as im:
                            w, h = im.size
                        scale = min(MAX_W / float(w or 1), MAX_H / float(h or 1), 1.0)
                        target_w, target_h = int(w * scale), int(h * scale)

                        xl_img = XLImage(img_path)
                        xl_img.width = target_w
                        xl_img.height = target_h
                        anchor = f"D{row_idx}"
                        ws2.add_image(xl_img, anchor)

                        # Ajuste de alto de fila (puntos). Aproximación: px * 0.75
                        ws2.row_dimensions[row_idx].height = max(
                            22, int(target_h * 0.75)
                        )
                    except Exception as e:
                        ws2.cell(
                            row=row_idx, column=4, value=f"(No se pudo incrustar: {e})"
                        )
                        ws2.row_dimensions[row_idx].height = 22
                else:
                    # No imagen compatible (pdf u otro)
                    ws2.cell(row=row_idx, column=4, value="(No es imagen o no existe)")
                    ws2.row_dimensions[row_idx].height = 22

                row_idx += 1

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    suffix = ""
    if estado:
        suffix += f"_{estado.lower()}"
    if sociedad_str:
        suffix += f"_{sociedad_str.lower()}"
    if desde_str or hasta_str:
        suf_d = desde_str or ""
        suf_h = hasta_str or ""
        suffix += f"_{suf_d}_a_{suf_h}"
    if include_images:
        suffix += "_con_imagenes"

    filename = f"bandeja_pagos{suffix}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
